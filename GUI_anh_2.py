from flask import Flask, request, render_template_string, send_from_directory, redirect, url_for
import os
import re
import shutil
import datetime

# Thêm vào dòng import đầu file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file
from flask import jsonify




CAMERA_NAMES = {

        "ch0": "Camera Trước",

        "ch1": "Camera Sau",

        "ch2": "Camera Trái",

        "ch3": "Camera Phải"

    }
app = Flask(__name__)

SAVE_DIR = "./received"
os.makedirs(SAVE_DIR, exist_ok=True)



@app.route("/api/stats/daily")
def api_daily_stats():
    """Trả về số ảnh theo ngày × camera"""
    all_days = get_all_days()
    from collections import defaultdict
    result = []
    all_cams = set()
    for day in all_days:
        day_dir = os.path.join(SAVE_DIR, day["date"])
        cam_data = {}
        for ch in sorted(os.listdir(day_dir)):
            ch_path = os.path.join(day_dir, ch)
            if not os.path.isdir(ch_path): continue
            count = len([f for f in os.listdir(ch_path) if is_image(f)])
            if count > 0:
                cam_data[ch] = count
                all_cams.add(ch)
        result.append({"date": day["date"], "cameras": cam_data})
    return jsonify({"days": result, "cameras": sorted(all_cams)})

@app.route("/api/stats/hourly/<date>")
def api_hourly_stats(date):
    """Trả về số ảnh theo giờ × camera của 1 ngày"""
    channel = request.args.get("ch", "all")
    images = get_images_by_day(date, channel)
    from collections import defaultdict
    hour_cam = defaultdict(lambda: defaultdict(int))
    all_cams = set()
    for img in images:
        if img["time"]:
            h = int(img["time"].split(":")[0])
            hour_cam[h][img["channel"]] += 1
            all_cams.add(img["channel"])
    hours = sorted(hour_cam.keys())
    result = [{"hour": f"{h:02d}:00", "cameras": dict(hour_cam[h])} for h in hours]
    return jsonify({"hours": result, "cameras": sorted(all_cams)})




@app.route("/stats")
def stats():
    all_days = get_all_days()
    
    # Tổng hợp số liệu toàn bộ
    total_images = sum(d["total"] for d in all_days)
    total_days = len(all_days)
    
    # Thống kê theo camera (toàn bộ)
    from collections import defaultdict
    cam_total = defaultdict(int)
    day_stats = []
    
    for day in all_days:
        day_dir = os.path.join(SAVE_DIR, day["date"])
        day_cam = {}
        day_count = 0
        
        for ch in sorted(os.listdir(day_dir)):
            ch_path = os.path.join(day_dir, ch)
            if not os.path.isdir(ch_path):
                continue
            count = len([f for f in os.listdir(ch_path) if is_image(f)])
            if count > 0:
                day_cam[ch] = count
                cam_total[ch] += count
                day_count += count
        
        # Giờ cao điểm trong ngày
        images = get_images_by_day(day["date"])
        hour_count = defaultdict(int)
        for img in images:
            if img["time"]:
                h = int(img["time"].split(":")[0])
                hour_count[h] += 1
        peak_hour = max(hour_count, key=hour_count.get) if hour_count else None
        peak_label = f"{peak_hour:02d}:00" if peak_hour is not None else "–"
        
        day_stats.append({
            "date": day["date"],
            "total": day_count,
            "cameras": day_cam,
            "peak_hour": peak_label,
            "peak_count": hour_count.get(peak_hour, 0) if peak_hour is not None else 0
        })
    
    all_cams = sorted(cam_total.keys())
    
    return render_template_string(
        STATS_HTML,
        day_stats=day_stats,
        cam_total=cam_total,
        all_cams=all_cams,
        total_images=total_images,
        total_days=total_days,
        CAMERA_NAMES=CAMERA_NAMES  # ← THÊM DÒNG NÀY
    )

@app.route("/export/<date>")
def export_excel(date):
    channel_filter = request.args.get("ch", "all")
    images = get_images_by_day(date, channel_filter)

    wb = Workbook()

    # ── Sheet 1: Danh sách ảnh chi tiết ──────────────────────────────────────

    ws1 = wb.active
    ws1.title = "Chi tiết vi phạm"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── HEADER: Tên công ty, tiêu đề, ngày ───────────────────────────────────
    COMPANY_NAME = "CÔNG TY CỔ PHẦN CÔNG NGHỆ IPX"   # ← sửa tên công ty ở đây
    REPORT_TITLE = "BÁO CÁO VI PHẠM CAMERA AN NINH"
    

    NUM_COLS = 6  # số cột của bảng dữ liệu

    # Dòng 1: Tên công ty
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    c1 = ws1.cell(row=1, column=1, value=COMPANY_NAME)
    c1.font = Font(bold=True, name="Arial", size=13, color="1F2937")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 26

    # Dòng 2: Tiêu đề báo cáo
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    c2 = ws1.cell(row=2, column=1, value=REPORT_TITLE)
    c2.font = Font(bold=True, name="Arial", size=15, color="1F2937")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 30

    # Dòng 3: Ngày tháng + camera lọc
    ch_label = f"Camera: {channel_filter.upper()}" if channel_filter != "all" else "Camera: Tất cả"
    ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS)
    c3 = ws1.cell(row=3, column=1,
                  value=f"Ngày: {date}    |    {ch_label}    |    Xuất lúc: {datetime.datetime.now().strftime('%H:%M %d/%m/%Y')}")
    c3.font = Font(name="Arial", size=10, italic=True, color="6B7280")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 20

    # Dòng 4: kẻ đường ngăn cách
    sep_fill = PatternFill("solid", fgColor="1F2937")
    for col in range(1, NUM_COLS + 1):
        ws1.cell(row=4, column=col).fill = sep_fill
    ws1.row_dimensions[4].height = 4

    # ── Header bảng dữ liệu (dòng 5) ─────────────────────────────────────────
    HEADER_ROW = 5
    headers = ["STT", "Ngày", "Giờ", "Camera", "Tên file", "Link ảnh"]
    col_widths = [6, 14, 12, 12, 50, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.row_dimensions[HEADER_ROW].height = 28

    alt_fill = PatternFill("solid", fgColor="F9FAFB")

    for i, img in enumerate(images, 1):
        row = i + HEADER_ROW   # ← dịch xuống sau header
        fill = alt_fill if i % 2 == 0 else PatternFill()
        host = request.host_url.rstrip("/")
        link = f"{host}{img['url']}"

        values = [i, date, img["time"], CAMERA_NAMES.get(img["channel"], img["channel"]), img["filename"]]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(
                horizontal="center" if col != 5 else "left",
                vertical="center"
            )
            cell.border = border
            if fill.fill_type:
                cell.fill = fill

        link_cell = ws1.cell(row=row, column=6, value="🔗 Xem ảnh")
        link_cell.hyperlink = link
        link_cell.font = Font(name="Arial", size=10, color="2563EB", underline="single")
        link_cell.alignment = Alignment(horizontal="center", vertical="center")
        link_cell.border = border
        if fill.fill_type:
            link_cell.fill = fill

        ws1.row_dimensions[row].height = 22
    # ws1 = wb.active
    # ws1.title = "Chi tiết vi phạm"

    # header_fill = PatternFill("solid", fgColor="1F2937")
    # header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    # center = Alignment(horizontal="center", vertical="center")
    # thin = Side(style="thin", color="D1D5DB")
    # border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # headers = ["STT", "Ngày", "Giờ", "Camera", "Tên file", "Link ảnh"]
    # col_widths = [6, 14, 12, 12, 50]

    # for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    #     cell = ws1.cell(row=1, column=col, value=h)
    #     cell.font = header_font
    #     cell.fill = header_fill
    #     cell.alignment = center
    #     cell.border = border
    #     ws1.column_dimensions[get_column_letter(col)].width = w

    # ws1.row_dimensions[1].height = 28

    # alt_fill = PatternFill("solid", fgColor="F9FAFB")

    # for i, img in enumerate(images, 1):
    #     row = i + 1
    #     fill = alt_fill if i % 2 == 0 else PatternFill()
    #     host = request.host_url.rstrip("/")
    #     link = f"{host}{img['url']}"

    #     values = [i, date, img["time"], img["channel"], img["filename"]]
    #     for col, val in enumerate(values, 1):
    #         cell = ws1.cell(row=row, column=col, value=val)
    #         cell.font = Font(name="Arial", size=10)
    #         cell.alignment = Alignment(
    #             horizontal="center" if col != 5 else "left",
    #             vertical="center"
    #         )
    #         cell.border = border
    #         if fill.fill_type:
    #             cell.fill = fill

    #     link_cell = ws1.cell(row=row, column=6, value="🔗 Xem ảnh")
    #     link_cell.hyperlink = link
    #     link_cell.font = Font(name="Arial", size=10, color="2563EB", underline="single")
    #     link_cell.alignment = Alignment(horizontal="center", vertical="center")
    #     link_cell.border = border
    #     if fill.fill_type:
    #         link_cell.fill = fill

    #     ws1.row_dimensions[row].height = 22

    # ── Sheet 2: Thống kê theo camera ────────────────────────────────────────

    # ── Sheet 2: Thống kê theo camera ────────────────────────────────────────
    ws2 = wb.create_sheet("Thống kê Camera")

    from collections import defaultdict
    cam_stats = defaultdict(int)
    for img in images:
        cam_stats[img["channel"]] += 1

    # Header
    for col in range(1, 4):
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    c = ws2.cell(row=1, column=1, value=COMPANY_NAME)
    c.font = Font(bold=True, name="Arial", size=13, color="1F2937")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    c = ws2.cell(row=2, column=1, value="THỐNG KÊ THEO CAMERA")
    c.font = Font(bold=True, name="Arial", size=15, color="1F2937")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 30

    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    c = ws2.cell(row=3, column=1,
                 value=f"Ngày: {date}    |    Xuất lúc: {datetime.datetime.now().strftime('%H:%M %d/%m/%Y')}")
    c.font = Font(name="Arial", size=10, italic=True, color="6B7280")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 20

    sep_fill = PatternFill("solid", fgColor="1F2937")
    for col in range(1, 4):
        ws2.cell(row=4, column=col).fill = sep_fill
    ws2.row_dimensions[4].height = 4

    # Header bảng (dòng 5)
    S2_HEADER_ROW = 5
    for col, h in enumerate(["Camera", "Số lượng ảnh", "Tỷ lệ (%)"], 1):
        cell = ws2.cell(row=S2_HEADER_ROW, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 16
    ws2.row_dimensions[S2_HEADER_ROW].height = 28

    for i, (ch, cnt) in enumerate(sorted(cam_stats.items()), 1):
        row = i + S2_HEADER_ROW
        ws2.cell(row=row, column=1, value=CAMERA_NAMES.get(ch, ch)).border = border
        ws2.cell(row=row, column=2, value=cnt).border = border
        pct_cell = ws2.cell(row=row, column=3,
                            value=f"=B{row}/B{len(cam_stats) + S2_HEADER_ROW + 1}")
        pct_cell.number_format = "0.0%"
        pct_cell.border = border
        for col in range(1, 4):
            ws2.cell(row=row, column=col).font = Font(name="Arial", size=10)
            ws2.cell(row=row, column=col).alignment = center

    total_row = len(cam_stats) + S2_HEADER_ROW + 1
    ws2.cell(row=total_row, column=1, value="Tổng").font = Font(bold=True, name="Arial")
    ws2.cell(row=total_row, column=2,
             value=f"=SUM(B{S2_HEADER_ROW+1}:B{total_row-1})").font = Font(bold=True, name="Arial")
    for col in range(1, 4):
        ws2.cell(row=total_row, column=col).border = border
        ws2.cell(row=total_row, column=col).alignment = center
        ws2.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="E5E7EB")
    # ws2 = wb.create_sheet("Thống kê Camera")

    # from collections import defaultdict
    # cam_stats = defaultdict(int)
    # for img in images:
    #     cam_stats[img["channel"]] += 1

    # for col, h in enumerate(["Camera", "Số lượng ảnh", "Tỷ lệ (%)"], 1):
    #     cell = ws2.cell(row=1, column=col, value=h)
    #     cell.font = header_font
    #     cell.fill = header_fill
    #     cell.alignment = center
    #     cell.border = border

    # ws2.column_dimensions["A"].width = 14
    # ws2.column_dimensions["B"].width = 18
    # ws2.column_dimensions["C"].width = 16

    # total = len(images)
    # for i, (ch, cnt) in enumerate(sorted(cam_stats.items()), 1):
    #     row = i + 1
    #     ws2.cell(row=row, column=1, value=ch).border = border
    #     ws2.cell(row=row, column=2, value=cnt).border = border
    #     pct_cell = ws2.cell(row=row, column=3,
    #                         value=f"=B{row}/B{len(cam_stats)+2}")
    #     pct_cell.number_format = "0.0%"
    #     pct_cell.border = border
    #     for col in range(1, 4):
    #         ws2.cell(row=row, column=col).font = Font(name="Arial", size=10)
    #         ws2.cell(row=row, column=col).alignment = center

    # total_row = len(cam_stats) + 2
    # ws2.cell(row=total_row, column=1, value="Tổng").font = Font(bold=True, name="Arial")
    # ws2.cell(row=total_row, column=2,
    #          value=f"=SUM(B2:B{total_row-1})").font = Font(bold=True, name="Arial")
    # for col in range(1, 4):
    #     ws2.cell(row=total_row, column=col).border = border
    #     ws2.cell(row=total_row, column=col).alignment = center
    #     ws2.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="E5E7EB")

    # ── Sheet 3: Phân tích khung giờ vi phạm ─────────────────────────────────

    # ── Sheet 3: Phân tích khung giờ vi phạm ─────────────────────────────────
    ws3 = wb.create_sheet("Khung giờ vi phạm")

    hour_cam = defaultdict(lambda: defaultdict(int))
    for img in images:
        if img["time"]:
            hour = int(img["time"].split(":")[0])
            hour_cam[hour][img["channel"]] += 1

    all_cams = sorted(set(img["channel"] for img in images))
    NUM_COLS_S3 = len(all_cams) + 2  # Khung giờ + cams + Tổng

    # Header
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS_S3)
    c = ws3.cell(row=1, column=1, value=COMPANY_NAME)
    c.font = Font(bold=True, name="Arial", size=13, color="1F2937")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS_S3)
    c = ws3.cell(row=2, column=1, value="PHÂN TÍCH KHUNG GIỜ VI PHẠM")
    c.font = Font(bold=True, name="Arial", size=15, color="1F2937")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 30

    ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS_S3)
    c = ws3.cell(row=3, column=1,
                 value=f"Ngày: {date}    |    Xuất lúc: {datetime.datetime.now().strftime('%H:%M %d/%m/%Y')}")
    c.font = Font(name="Arial", size=10, italic=True, color="6B7280")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[3].height = 20

    sep_fill = PatternFill("solid", fgColor="1F2937")
    for col in range(1, NUM_COLS_S3 + 1):
        ws3.cell(row=4, column=col).fill = sep_fill
    ws3.row_dimensions[4].height = 4

    # Header bảng (dòng 5)
    S3_HEADER_ROW = 5
    ws3.cell(row=S3_HEADER_ROW, column=1, value="Khung giờ").font = header_font
    ws3.cell(row=S3_HEADER_ROW, column=1).fill = header_fill
    ws3.cell(row=S3_HEADER_ROW, column=1).alignment = center
    ws3.cell(row=S3_HEADER_ROW, column=1).border = border
    ws3.column_dimensions["A"].width = 16
    ws3.row_dimensions[S3_HEADER_ROW].height = 28

    for j, cam in enumerate(all_cams, 2):
        cell = ws3.cell(row=S3_HEADER_ROW, column=j, value=CAMERA_NAMES.get(cam, cam))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws3.column_dimensions[get_column_letter(j)].width = 14

    total_col = len(all_cams) + 2
    ws3.cell(row=S3_HEADER_ROW, column=total_col, value="Tổng").font = header_font
    ws3.cell(row=S3_HEADER_ROW, column=total_col).fill = header_fill
    ws3.cell(row=S3_HEADER_ROW, column=total_col).alignment = center
    ws3.cell(row=S3_HEADER_ROW, column=total_col).border = border
    ws3.column_dimensions[get_column_letter(total_col)].width = 10

    high_fill = PatternFill("solid", fgColor="FEE2E2")
    mid_fill  = PatternFill("solid", fgColor="FEF9C3")
    zero_font = Font(name="Arial", size=10, color="9CA3AF")

    sorted_hours = sorted(hour_cam.keys())
    for i, hour in enumerate(sorted_hours, 1):
        row = i + S3_HEADER_ROW
        label = f"{hour:02d}:00 – {hour:02d}:59"
        ws3.cell(row=row, column=1, value=label).border = border
        ws3.cell(row=row, column=1).alignment = center
        ws3.cell(row=row, column=1).font = Font(name="Arial", size=10)

        row_total = 0
        for j, cam in enumerate(all_cams, 2):
            cnt = hour_cam[hour].get(cam, 0)
            row_total += cnt
            cell = ws3.cell(row=row, column=j, value=cnt if cnt else "–")
            cell.border = border
            cell.alignment = center
            cell.font = zero_font if cnt == 0 else Font(name="Arial", size=10, bold=cnt >= 5)

        fill = high_fill if row_total >= 10 else (mid_fill if row_total >= 3 else PatternFill())
        if fill.fill_type:
            for j in range(1, total_col + 1):
                ws3.cell(row=row, column=j).fill = fill

        total_cell = ws3.cell(row=row, column=total_col, value=row_total)
        total_cell.border = border
        total_cell.alignment = center
        total_cell.font = Font(name="Arial", size=10, bold=True)

    # Chú thích màu
    note_row = len(sorted_hours) + S3_HEADER_ROW + 2
    ws3.cell(row=note_row, column=1, value="Chú thích:").font = Font(bold=True, name="Arial")
    ws3.cell(row=note_row+1, column=1, value="Nền đỏ").fill = high_fill
    ws3.cell(row=note_row+1, column=2, value="≥ 10 ảnh – Khung giờ vi phạm nhiều").font = Font(name="Arial", size=10)
    ws3.cell(row=note_row+2, column=1, value="Nền vàng").fill = mid_fill
    ws3.cell(row=note_row+2, column=2, value="3–9 ảnh – Khung giờ vi phạm trung bình").font = Font(name="Arial", size=10)
    # ws3 = wb.create_sheet("Khung giờ vi phạm")

    # hour_cam = defaultdict(lambda: defaultdict(int))  # hour -> cam -> count
    # for img in images:
    #     if img["time"]:
    #         hour = int(img["time"].split(":")[0])
    #         hour_cam[hour][img["channel"]] += 1

    # all_cams = sorted(set(img["channel"] for img in images))

    # # Header row
    # ws3.cell(row=1, column=1, value="Khung giờ").font = header_font
    # ws3.cell(row=1, column=1).fill = header_fill
    # ws3.cell(row=1, column=1).alignment = center
    # ws3.cell(row=1, column=1).border = border
    # ws3.column_dimensions["A"].width = 16

    # for j, cam in enumerate(all_cams, 2):
    #     cell = ws3.cell(row=1, column=j, value=cam)
    #     cell.font = header_font
    #     cell.fill = header_fill
    #     cell.alignment = center
    #     cell.border = border
    #     ws3.column_dimensions[get_column_letter(j)].width = 14

    # total_col = len(all_cams) + 2
    # ws3.cell(row=1, column=total_col, value="Tổng").font = header_font
    # ws3.cell(row=1, column=total_col).fill = header_fill
    # ws3.cell(row=1, column=total_col).alignment = center
    # ws3.cell(row=1, column=total_col).border = border
    # ws3.column_dimensions[get_column_letter(total_col)].width = 10

    # high_fill = PatternFill("solid", fgColor="FEE2E2")   # đỏ nhạt = nhiều vi phạm
    # mid_fill  = PatternFill("solid", fgColor="FEF9C3")   # vàng nhạt = trung bình
    # zero_font = Font(name="Arial", size=10, color="9CA3AF")

    # sorted_hours = sorted(hour_cam.keys())
    # for i, hour in enumerate(sorted_hours, 2):
    #     label = f"{hour:02d}:00 – {hour:02d}:59"
    #     ws3.cell(row=i, column=1, value=label).border = border
    #     ws3.cell(row=i, column=1).alignment = center
    #     ws3.cell(row=i, column=1).font = Font(name="Arial", size=10)

    #     row_total = 0
    #     for j, cam in enumerate(all_cams, 2):
    #         cnt = hour_cam[hour].get(cam, 0)
    #         row_total += cnt
    #         cell = ws3.cell(row=i, column=j, value=cnt if cnt else "–")
    #         cell.border = border
    #         cell.alignment = center
    #         cell.font = zero_font if cnt == 0 else Font(name="Arial", size=10, bold=cnt >= 5)

    #     # Tô màu cả hàng theo mức vi phạm
    #     fill = high_fill if row_total >= 10 else (mid_fill if row_total >= 3 else PatternFill())
    #     if fill.fill_type:
    #         for j in range(1, total_col + 1):
    #             ws3.cell(row=i, column=j).fill = fill

    #     total_cell = ws3.cell(row=i, column=total_col, value=row_total)
    #     total_cell.border = border
    #     total_cell.alignment = center
    #     total_cell.font = Font(name="Arial", size=10, bold=True)

    # # Chú thích màu
    # note_row = len(sorted_hours) + 3
    # ws3.cell(row=note_row, column=1, value="Chú thích:").font = Font(bold=True, name="Arial")
    # ws3.cell(row=note_row+1, column=1, value="Nền đỏ").fill = high_fill
    # ws3.cell(row=note_row+1, column=2, value="≥ 10 ảnh – Khung giờ vi phạm nhiều").font = Font(name="Arial", size=10)
    # ws3.cell(row=note_row+2, column=1, value="Nền vàng").fill = mid_fill
    # ws3.cell(row=note_row+2, column=2, value="3–9 ảnh – Khung giờ vi phạm trung bình").font = Font(name="Arial", size=10)

    # ── Xuất file ─────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_date = date.replace("-", "")
    filename = f"vi_pham_{safe_date}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def is_image(name):
    return name.lower().endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp"))


def parse_image_name(filename):
    """
    Ví dụ:
    20260425_154922_722699_ch1_7781031.jpg

    Tách ra:
    date    = 2026-04-25
    time    = 15:49:22
    channel = ch1
    """
    m = re.search(r"(\d{8})_(\d{6})_\d+_(ch\d+)_", filename)
    if not m:
        return None, None, None

    raw_date = m.group(1)
    raw_time = m.group(2)
    channel = m.group(3)

    date = f"{raw_date[0:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
    time = f"{raw_time[0:2]}:{raw_time[2:4]}:{raw_time[4:6]}"

    return date, time, channel


def safe_move(src, dst_dir, filename):
    os.makedirs(dst_dir, exist_ok=True)

    dst = os.path.join(dst_dir, filename)

    if not os.path.exists(dst):
        shutil.move(src, dst)
        return

    name, ext = os.path.splitext(filename)
    new_name = f"{name}_{datetime.datetime.now().strftime('%H%M%S_%f')}{ext}"
    shutil.move(src, os.path.join(dst_dir, new_name))


def migrate_old_images():
    """
    Tự động đọc ảnh cũ đang nằm trực tiếp trong received/
    rồi chuyển vào đúng ngày/kênh.
    """
    for name in os.listdir(SAVE_DIR):
        path = os.path.join(SAVE_DIR, name)

        if not os.path.isfile(path):
            continue

        if not is_image(name):
            continue

        date, _, channel = parse_image_name(name)

        if date and channel:
            target_dir = os.path.join(SAVE_DIR, date, channel)
        else:
            target_dir = os.path.join(SAVE_DIR, "_anh_cu", "unknown")

        safe_move(path, target_dir, name)


def get_all_days():
    migrate_old_images()

    days = []

    for day in sorted(os.listdir(SAVE_DIR), reverse=True):
        day_path = os.path.join(SAVE_DIR, day)

        if not os.path.isdir(day_path):
            continue

        total = 0
        cover = None
        cover_ch = None
        channels = {}

        for ch in sorted(os.listdir(day_path)):
            ch_path = os.path.join(day_path, ch)

            if not os.path.isdir(ch_path):
                continue

            imgs = [
                img for img in sorted(os.listdir(ch_path), reverse=True)
                if is_image(img)
            ]

            if imgs:
                channels[ch] = len(imgs)
                total += len(imgs)

                if cover is None:
                    cover = imgs[0]
                    cover_ch = ch

        if total > 0:
            days.append({
                "date": day,
                "total": total,
                "cover": cover,
                "cover_ch": cover_ch,
                "channels": channels
            })

    return days


def get_images_by_day(date, channel="all"):
    migrate_old_images()

    day_dir = os.path.join(SAVE_DIR, date)
    result = []

    if not os.path.exists(day_dir):
        return []

    channels = sorted(os.listdir(day_dir))

    for ch in channels:
        ch_dir = os.path.join(day_dir, ch)

        if not os.path.isdir(ch_dir):
            continue

        if channel != "all" and ch != channel:
            continue

        for img in sorted(os.listdir(ch_dir), reverse=True):
            if not is_image(img):
                continue

            _, t, parsed_ch = parse_image_name(img)

            result.append({
                "filename": img,
                "channel": parsed_ch or ch,
                "time": t or "",
                "url": url_for("get_image", date=date, channel=ch, filename=img)
            })

    return result

@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("image")
    if not f:
        return "no image", 400

    original_name = f.filename  # Giữ nguyên tên từ C++ gửi lên

    # Parse date/time/channel từ tên file gốc
    date, time_str, channel = parse_image_name(original_name)

    if date and channel:
        save_dir = os.path.join(SAVE_DIR, date, channel)
        os.makedirs(save_dir, exist_ok=True)

        dst = os.path.join(save_dir, original_name)
        if os.path.exists(dst):
            # Tránh ghi đè: thêm microsecond vào tên
            name, ext = os.path.splitext(original_name)
            original_name = f"{name}_{datetime.datetime.now().strftime('%f')}{ext}"

        f.save(os.path.join(save_dir, original_name))
        print(f"[RECV] {date}/{channel}/{original_name}  time={time_str}")
    else:
        # Fallback nếu tên không đúng format
        now = datetime.datetime.now()
        date = now.strftime("%Y-%m-%d")
        m = re.search(r"(ch\d+)", f.filename)
        channel = m.group(1) if m else "unknown"
        save_dir = os.path.join(SAVE_DIR, date, channel)
        os.makedirs(save_dir, exist_ok=True)
        f.save(os.path.join(save_dir, f"{now.strftime('%Y%m%d_%H%M%S_%f')}_{f.filename}"))
        print(f"[RECV fallback] {date}/{channel}/{f.filename}")

    return "ok", 200


# @app.route("/upload", methods=["POST"])
# def upload():
#     f = request.files.get("image")
#     if not f:
#         return "no image", 400

#     now = datetime.datetime.now()

#     # giữ tên giống code cũ của bạn
#     fname = f"{now.strftime('%Y%m%d_%H%M%S_%f')}_{f.filename}"

#     # chia ngày theo thời điểm server nhận
#     date = now.strftime("%Y-%m-%d")

#     # lấy channel từ filename gốc, ví dụ ch0_13315.jpg
#     m = re.search(r"(ch\d+)", f.filename)
#     channel = m.group(1) if m else "unknown"

#     save_dir = os.path.join(SAVE_DIR, date, channel)
#     os.makedirs(save_dir, exist_ok=True)

#     f.save(os.path.join(save_dir, fname))

#     print(f"[RECV] {date}/{channel}/{fname}")
#     return "ok", 200


@app.route("/")
def index():
    days = get_all_days()
    return render_template_string(INDEX_HTML, days=days, CAMERA_NAMES=CAMERA_NAMES)


# @app.route("/day/<date>")
# def day_view(date):
#     channel = request.args.get("ch", "all")
#     images = get_images_by_day(date, channel)

#     available_channels = []

#     day_dir = os.path.join(SAVE_DIR, date)
#     if os.path.exists(day_dir):
#         available_channels = [
#             ch for ch in sorted(os.listdir(day_dir))
#             if os.path.isdir(os.path.join(day_dir, ch))
#         ]

#     return render_template_string(
#         DAY_HTML,
#         date=date,
#         images=images,
#         channel=channel,
#         available_channels=available_channels
#     )

@app.route("/day/<date>")
def day_view(date):
    channel = request.args.get("ch", "all")
    images = get_images_by_day(date, channel)

    available_channels = []
    day_dir = os.path.join(SAVE_DIR, date)
    if os.path.exists(day_dir):
        available_channels = [
            ch for ch in sorted(os.listdir(day_dir))
            if os.path.isdir(os.path.join(day_dir, ch))
        ]

    # Tính thống kê cho KPI
    from collections import defaultdict, Counter
    cam_counts = Counter(img["channel"] for img in images)
    hour_counter = Counter(
        int(img["time"].split(":")[0])
        for img in images if img["time"]
    )
    peak_hour = max(hour_counter, key=hour_counter.get) if hour_counter else None
    stats = {
        "total": len(images),
        "cam_counts": dict(cam_counts),
        "peak_hour": f"{peak_hour:02d}:00" if peak_hour is not None else None,
        "peak_count": hour_counter.get(peak_hour, 0) if peak_hour else 0,
    }

    return render_template_string(
        DAY_HTML,
        date=date,
        images=images,
        channel=channel,
        available_channels=available_channels,
        stats=stats,
        CAMERA_NAMES=CAMERA_NAMES  # THÊM DÒNG NÀY VÀO ĐÂY
    )


@app.route("/image/<date>/<channel>/<filename>")
def get_image(date, channel, filename):
    return send_from_directory(os.path.join(SAVE_DIR, date, channel), filename)


@app.route("/delete/<date>/<channel>/<filename>", methods=["POST"])
def delete_image(date, channel, filename):
    path = os.path.join(SAVE_DIR, date, channel, filename)

    if os.path.exists(path):
        os.remove(path)

    return redirect(url_for("day_view", date=date, ch=request.form.get("current_ch", "all")))


INDEX_HTML = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Quản lý ảnh cảnh báo</title>

    <style>
        body {
            margin: 0;
            font-family: Arial, sans-serif;
            background: #f3f4f6;
            color: #111827;
        }

        header {
            background: #111827;
            color: white;
            padding: 22px 32px;
        }

        h1 {
            margin: 0;
            font-size: 27px;
        }

        .sub {
            margin-top: 6px;
            color: #cbd5e1;
            font-size: 14px;
        }

        .container {
            padding: 28px;
        }

        .grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
            gap: 22px;
        }

        .card {
            background: white;
            border-radius: 18px;
            overflow: hidden;
            box-shadow: 0 8px 22px rgba(0,0,0,0.08);
            text-decoration: none;
            color: #111827;
            transition: 0.18s;
        }

        .card:hover {
            transform: translateY(-4px);
            box-shadow: 0 14px 32px rgba(0,0,0,0.16);
        }

        .thumb-wrap {
            position: relative;
            height: 190px;
            background: #e5e7eb;
        }

        .thumb-wrap img {
            width: 100%;
            height: 100%;
            object-fit: cover;
        }

        .total-badge {
            position: absolute;
            top: 12px;
            right: 12px;
            background: rgba(17,24,39,0.88);
            color: white;
            padding: 7px 11px;
            border-radius: 999px;
            font-size: 13px;
        }

        .info {
            padding: 16px;
        }

        .date {
            font-size: 21px;
            font-weight: 700;
        }

        .channels {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-top: 13px;
        }

        .chip {
            background: #eef2ff;
            color: #3730a3;
            padding: 6px 10px;
            border-radius: 999px;
            font-size: 13px;
        }

        .empty {
            padding: 40px;
            background: white;
            border-radius: 14px;
            text-align: center;
            color: #6b7280;
        }
        .popup-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.55);z-index:999;align-items:center;justify-content:center}
        .popup-overlay.show{display:flex}
        .popup-box{background:white;border-radius:18px;width:680px;max-width:95vw;padding:24px 28px;box-shadow:0 20px 60px rgba(0,0,0,0.25)}
        .popup-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:16px}
        .popup-header h2{margin:0;font-size:18px;font-weight:700}
        .close-btn{background:#f3f4f6;border:none;border-radius:10px;padding:8px 14px;cursor:pointer;font-size:14px}
        .legend-row{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:12px;font-size:13px;color:#6b7280}
        .legend-dot{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:4px}
        }
    </style>
</head>
<body>

    <header style="display: flex; justify-content: space-between; align-items: center; background: #111827; color: white; padding: 22px 32px;">
        
        <div>
            <h1 style="margin: 0; font-size: 27px;">📸 Quản lý ảnh cảnh báo</h1>
            <div class="sub" style="margin-top: 6px; color: #cbd5e1; font-size: 14px;">Tự chia theo ngày và camera</div>
        </div>

        <div style="display: flex; gap: 10px;">
            <button onclick="openDayChart()" style="background: #2563eb; color: white; padding: 11px 18px; border-radius: 10px; border: none; cursor: pointer;">
                📊 Biểu đồ theo ngày
            </button>
            <a href="/stats" style="background: #4b5563; color: white; padding: 11px 18px; border-radius: 10px; text-decoration: none;">
                📊 Thống kê
            </a>
        </div>

    </header>

    <div class="container">
        {% if days %}
        <div class="grid">
            {% for day in days %}
            <a class="card" href="/day/{{ day.date }}">
                <div class="thumb-wrap">
                    <img src="/image/{{ day.date }}/{{ day.cover_ch }}/{{ day.cover }}">
                    <div class="total-badge">{{ day.total }} ảnh</div>
                </div>

                <div class="info">
                    <div class="date">
                        {% if day.date == "_anh_cu" %}
                            📂 Ảnh cũ
                        {% else %}
                            📅 {{ day.date }}
                        {% endif %}
                    </div>

                    <div class="channels">
                        {% for ch, count in day.channels.items() %}
                        <span class="chip">{{ CAMERA_NAMES.get(ch, ch) }}: {{ count }}</span>
                        {% endfor %}
                    </div>
                </div>
            </a>
            {% endfor %}
        </div>
        {% else %}
        <div class="empty">
            Chưa có ảnh nào được upload.
        </div>
        {% endif %}
    </div>


<div class="popup-overlay" id="dayOverlay" onclick="if(event.target===this)closeDayChart()">
  <div class="popup-box">
    <div class="popup-header">
      <h2>📊 Số ảnh vi phạm theo ngày</h2>
      <button class="close-btn" onclick="closeDayChart()">✕ Đóng</button>
    </div>
    <div class="legend-row" id="dayLegend"></div>
    <div style="position:relative;width:100%;height:300px">
      <canvas id="dayChart" role="img" aria-label="Biểu đồ số ảnh vi phạm theo ngày"></canvas>
    </div>
  </div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const CAM_NAMES = {{ CAMERA_NAMES | tojson }};
const CAM_COLORS = ['#378ADD','#1D9E75','#D85A30','#888780','#D4537E','#BA7517'];
let dayChartInst = null;

async function openDayChart() {
  document.getElementById('dayOverlay').classList.add('show');
  if (dayChartInst) return;  // đã load rồi thì không load lại
  const res = await fetch('/api/stats/daily');
  const data = await res.json();
  const labels = data.days.map(d => d.date);
  const datasets = data.cameras.map((cam, i) => ({
    label: CAM_NAMES[cam] || cam,
    data: data.days.map(d => d.cameras[cam] || 0),
    backgroundColor: CAM_COLORS[i % CAM_COLORS.length],
    borderRadius: 4
  }));
  buildLegend('dayLegend', datasets);
  dayChartInst = new Chart(document.getElementById('dayChart'), {
    type: 'bar',
    data: { labels, datasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { display: false } },
      scales: {
        x: { stacked: true, ticks: { font:{size:11}, autoSkip: false, maxRotation: 45 }, grid:{display:false} },
        y: { stacked: true, ticks: { font:{size:11} }, grid:{color:'rgba(0,0,0,0.06)'} }
      }
    }
  });
}
function closeDayChart() { document.getElementById('dayOverlay').classList.remove('show'); }
function buildLegend(id, datasets) {
  document.getElementById(id).innerHTML = datasets.map((ds,i) =>
    `<span><span class="legend-dot" style="background:${ds.backgroundColor}"></span>${ds.label}</span>`
  ).join('');
}
</script>

<footer style="
    position: fixed;
    bottom: 0;
    left: 0;
    width: 100%;
    background: rgba(255, 255, 255, 0.95);
    backdrop-filter: blur(5px);
    border-top: 1px solid #e5e7eb;
    padding: 12px 0;
    z-index: 9999;
    box-shadow: 0 -2px 10px rgba(0,0,0,0.05);
">
    <div style="
        max-width: 1200px;
        margin: 0 auto;
        padding: 0 32px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    ">
        <div style="font-size: 13px; color: #4b5563;">
            <strong>Công ty cổ phần công nghệ IPX</strong> | 📍 32/19 Đường 494, phường Tăng Nhơn Phú, TP.HCM
        </div>

        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="font-size: 13px; color: #4b5563;">
                ✉️ ipx@gmail.com
            </div>
            <div style="
                background: #1d4ed8;
                color: white;
                padding: 6px 15px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
            ">
                Hotline: 0983 432 188
            </div>
        </div>
    </div>
</footer>

<div style="height: 70px;"></div>

</body>
</html>
"""

DAY_HTML = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Ảnh {{ date }}</title>
    <style>
        body { margin: 0; font-family: Arial, sans-serif; background: #f3f4f6; color: #111827; }
        header { background: #111827; color: white; padding: 20px 32px; display: flex; justify-content: space-between; gap: 20px; align-items: center; }
        h2 { margin: 0; font-size: 24px; }
        .meta { margin-top: 6px; color: #cbd5e1; font-size: 14px; }
        a.back { color: white; text-decoration: none; font-size: 15px; white-space: nowrap; }
        .container { padding: 24px 28px 34px; }
        .filters { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 22px; }
        .filter { text-decoration: none; background: white; color: #111827; padding: 10px 15px; border-radius: 999px; box-shadow: 0 3px 12px rgba(0,0,0,0.06); font-size: 14px; }
        .filter.active { background: #2563eb; color: white; }
        .grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(270px, 1fr)); gap: 22px; }
        .card { background: white; border-radius: 18px; overflow: hidden; box-shadow: 0 8px 22px rgba(0,0,0,0.08); }
        .img-wrap { position: relative; height: 210px; background: #e5e7eb; }
        .img-wrap img { width: 100%; height: 100%; object-fit: cover; cursor: pointer; }
        .ch-badge { position: absolute; top: 10px; left: 10px; background: rgba(37,99,235,0.95); color: white; padding: 6px 10px; border-radius: 999px; font-size: 13px; font-weight: 700; }
        .time-badge { position: absolute; top: 10px; right: 10px; background: rgba(17,24,39,0.88); color: white; padding: 6px 10px; border-radius: 999px; font-size: 13px; }
        .info { padding: 13px 14px 14px; }
        .actions { display: flex; gap: 8px; }
        .btn { flex: 1; border: none; padding: 10px; border-radius: 10px; cursor: pointer; text-align: center; text-decoration: none; font-size: 14px; }
        .download { background: #2563eb; color: white; }
        .delete { background: #dc2626; color: white; }
        .empty { padding: 38px; background: white; border-radius: 14px; text-align: center; color: #6b7280; }
        .modal { display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.88); justify-content: center; align-items: center; z-index: 999; }
        .modal.show { display: flex; }
        .modal img { max-width: 94vw; max-height: 92vh; border-radius: 12px; }
        .popup-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.55);z-index:999;align-items:center;justify-content:center}
        .popup-overlay.show{display:flex}
        .popup-box{background:white;border-radius:18px;width:720px;max-width:95vw;padding:24px 28px;box-shadow:0 20px 60px rgba(0,0,0,0.25)}
        .popup-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:16px}
        .popup-header h2{margin:0;font-size:18px;font-weight:700}
        .close-btn{background:#f3f4f6;border:none;border-radius:10px;padding:8px 14px;cursor:pointer;font-size:14px}
        .legend-row{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:12px;font-size:13px;color:#6b7280}
        .legend-dot{width:10px;height:10px;border-radius:2px;display:inline-block;margin-right:4px}
    </style>


    <div class="popup-overlay" id="hourOverlay" onclick="if(event.target===this)closeHourChart()">
        <div class="popup-box">
            <div class="popup-header">
            <h2>📈 Phân tích giờ — {{ date }}</h2>
            <button class="close-btn" onclick="closeHourChart()">✕ Đóng</button>
            </div>
            <div class="legend-row" id="hourLegend"></div>
            <div style="position:relative;width:100%;height:300px">
            <canvas id="hourChart" role="img" aria-label="Biểu đồ số ảnh theo giờ"></canvas>
            </div>
        </div>
    </div>

        <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
        <script>
        const CAM_NAMES = {{ CAMERA_NAMES | tojson }};
        const H_COLORS = ['#378ADD','#1D9E75','#D85A30','#888780','#D4537E','#BA7517'];
        let hourChartInst = null;

        async function openHourChart() {
        document.getElementById('hourOverlay').classList.add('show');
        if (hourChartInst) return;
        const res = await fetch('/api/stats/hourly/{{ date }}?ch={{ channel }}');
        const data = await res.json();
        const labels = data.hours.map(h => h.hour);
        const datasets = data.cameras.map((cam, i) => ({
            label: CAM_NAMES[cam] || cam,
            data: data.hours.map(h => h.cameras[cam] || 0),
            backgroundColor: H_COLORS[i % H_COLORS.length],
            borderRadius: 4
        }));
        document.getElementById('hourLegend').innerHTML = datasets.map(ds =>
            `<span><span class="legend-dot" style="background:${ds.backgroundColor}"></span>${ds.label}</span>`
        ).join('');
        hourChartInst = new Chart(document.getElementById('hourChart'), {
            type: 'bar',
            data: { labels, datasets },
            options: {
            responsive: true, maintainAspectRatio: false,
            plugins: { legend: { display: false } },
            scales: {
                x: { stacked: true, ticks:{font:{size:11}, autoSkip:false, maxRotation:0}, grid:{display:false} },
                y: { stacked: true, ticks:{font:{size:11}}, grid:{color:'rgba(0,0,0,0.06)'} }
            }
            }
        });
        }
        function closeHourChart() { document.getElementById('hourOverlay').classList.remove('show'); }
        </script>

</head>
<body>

<header>
    <div>
        <h2>
            {% if date == "_anh_cu" %}📂 Ảnh cũ
            {% else %}📅 Ngày {{ date }}{% endif %}
        </h2>
        <div class="meta">{{ images|length }} ảnh đang hiển thị</div>
    </div>
    <div style="display:flex; gap:12px; align-items:center;">
        <button onclick="openHourChart()"
            style="background:#7c3aed; color:white; padding:11px 18px; border-radius:10px; border:none; font-size:14px; cursor:pointer; white-space:nowrap;">
            📈 Biểu đồ theo giờ
        </button>

        <a href="/export/{{ date }}?ch={{ channel }}"
        style="background:#16a34a; color:white; padding:11px 18px; border-radius:10px; text-decoration:none; font-size:14px; white-space:nowrap;">
            📊 Xuất Excel
        </a>

        <a class="back" href="/" style="color:white; text-decoration:none; font-size:15px; white-space:nowrap;">
            ← Quay lại
        </a>
    </div>

</header>

<div class="container">

    <div class="filters">
        <a class="filter {% if channel == 'all' %}active{% endif %}" href="/day/{{ date }}?ch=all">Tất cả</a>
        {% for ch in available_channels %}
        <a class="filter {% if channel == ch %}active{% endif %}" href="/day/{{ date }}?ch={{ ch }}">
            {{ CAMERA_NAMES.get(ch, ch) }}
        </a>
        {% endfor %}
    </div>
    <!-- KPI cards — dùng stats từ Python, không tính trong template -->
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:14px;margin-bottom:24px;">

        <div style="background:white;border-radius:14px;padding:18px 20px;box-shadow:0 4px 14px rgba(0,0,0,0.07);">
            <div style="font-size:12px;color:#6b7280;margin-bottom:6px;">Tổng ảnh</div>
            <div style="font-size:30px;font-weight:700;">{{ stats.total }}</div>
            <div style="font-size:12px;color:#9ca3af;margin-top:4px;">ảnh vi phạm</div>
        </div>

        <div style="background:white;border-radius:14px;padding:18px 20px;box-shadow:0 4px 14px rgba(0,0,0,0.07);">
            <div style="font-size:12px;color:#6b7280;margin-bottom:6px;">Số camera</div>
            <div style="font-size:30px;font-weight:700;">{{ stats.cam_counts|length }}</div>
            <div style="font-size:12px;color:#9ca3af;margin-top:4px;">camera ghi nhận</div>
        </div>

        {% if stats.peak_hour %}
        <div style="background:white;border-radius:14px;padding:18px 20px;box-shadow:0 4px 14px rgba(0,0,0,0.07);">
            <div style="font-size:12px;color:#6b7280;margin-bottom:6px;">Giờ cao điểm</div>
            <div style="font-size:30px;font-weight:700;">{{ stats.peak_hour }}</div>
            <div style="font-size:12px;color:#9ca3af;margin-top:4px;">{{ stats.peak_count }} ảnh trong giờ đó</div>
        </div>
        {% endif %}

        {% for cam, cnt in stats.cam_counts.items() %}
        <div style="background:white;border-radius:14px;padding:18px 20px;box-shadow:0 4px 14px rgba(0,0,0,0.07);">
            <div style="font-size:12px;color:#6b7280;margin-bottom:6px;">{{ CAMERA_NAMES.get(cam, cam) }}</div>
            <div style="font-size:30px;font-weight:700;color:#2563eb;">{{ cnt }}</div>
            <div style="background:#e5e7eb;border-radius:999px;height:6px;margin-top:8px;">
                <div style="background:#2563eb;height:100%;border-radius:999px;width:{{ (cnt/stats.total*100)|round }}%;"></div>
            </div>
            <div style="font-size:11px;color:#9ca3af;margin-top:4px;">{{ (cnt/stats.total*100)|round(1) }}% tổng ảnh</div>
        </div>
        {% endfor %}

    </div>

    {% if images %}
    <div class="grid">
        {% for img in images %}
        <div class="card">
            <div class="img-wrap">
                <img src="{{ img.url }}" onclick="openModal(this.src)">
                <div class="ch-badge">{{ CAMERA_NAMES.get(img.channel, img.channel) }}</div>
                {% if img.time %}
                <div class="time-badge">{{ img.time }}</div>
                {% endif %}
            </div>
            <div class="info">
                <a href="{{ img.url }}" target="_blank"
                   style="display:block;font-size:13px;color:#2563eb;word-break:break-all;margin-bottom:11px;text-decoration:none;"
                   title="Mở ảnh trong tab mới">
                    🔗 {{ img.filename }}
                </a>
                <div class="actions">
                    <a class="btn download" href="{{ img.url }}" download>Tải</a>
                    <form method="POST"
                          action="/delete/{{ date }}/{{ img.channel }}/{{ img.filename }}"
                          style="flex:1;"
                          onsubmit="return confirm('Xoá ảnh này?')">
                        <input type="hidden" name="current_ch" value="{{ channel }}">
                        <button class="btn delete" type="submit" style="width:100%;">Xoá</button>
                    </form>
                </div>
            </div>
        </div>
        {% endfor %}
    </div>
    {% else %}
    <div class="empty">Không có ảnh trong mục này.</div>
    {% endif %}

</div>

<div id="modal" class="modal" onclick="closeModal()">
    <img id="modalImg">
</div>

<script>
function openModal(src) {
    document.getElementById("modalImg").src = src;
    document.getElementById("modal").classList.add("show");
}
function closeModal() {
    document.getElementById("modal").classList.remove("show");
}
</script>

<footer style="
    position: fixed;
    bottom: 0;
    left: 0;
    width: 100%;
    background: rgba(255, 255, 255, 0.95);
    backdrop-filter: blur(5px);
    border-top: 1px solid #e5e7eb;
    padding: 12px 0;
    z-index: 9999;
    box-shadow: 0 -2px 10px rgba(0,0,0,0.05);
">
    <div style="
        max-width: 1200px;
        margin: 0 auto;
        padding: 0 32px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    ">
        <div style="font-size: 13px; color: #4b5563;">
            <strong>Công ty cổ phần công nghệ IPX IPX</strong> | 📍 32/19 Đường 494, phường Tăng Nhơn Phú, TP.HCM
        </div>

        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="font-size: 13px; color: #4b5563;">
                ✉️ ipx@gmail.com
            </div>
            <div style="
                background: #1d4ed8;
                color: white;
                padding: 6px 15px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
            ">
                Hotline: 0983 432 188
            </div>
        </div>
    </div>
</footer>


</body>
</html>
"""

STATS_HTML = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Thống kê vi phạm</title>
    <style>
        * { box-sizing: border-box; }
        body { margin: 0; font-family: Arial, sans-serif; background: #f3f4f6; color: #111827; }

        header {
            background: #111827; color: white;
            padding: 20px 32px;
            display: flex; justify-content: space-between; align-items: center;
        }
        header h1 { margin: 0; font-size: 24px; }
        a.back { color: #cbd5e1; text-decoration: none; font-size: 14px; }
        a.back:hover { color: white; }

        .container { padding: 28px; }

        /* KPI cards */
        .kpi-row {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
            gap: 16px;
            margin-bottom: 28px;
        }
        .kpi {
            background: white; border-radius: 14px;
            padding: 20px 22px;
            box-shadow: 0 4px 14px rgba(0,0,0,0.07);
        }
        .kpi .label { font-size: 13px; color: #6b7280; margin-bottom: 8px; }
        .kpi .value { font-size: 32px; font-weight: 700; color: #111827; }
        .kpi .sub { font-size: 12px; color: #9ca3af; margin-top: 4px; }

        /* Section */
        .section {
            background: white; border-radius: 16px;
            padding: 24px 26px;
            box-shadow: 0 4px 14px rgba(0,0,0,0.07);
            margin-bottom: 24px;
        }
        .section h2 { margin: 0 0 18px; font-size: 17px; color: #1f2937; }

        /* Table */
        table { width: 100%; border-collapse: collapse; font-size: 14px; }
        thead tr { background: #1f2937; color: white; }
        thead th { padding: 11px 14px; text-align: center; font-weight: 600; }
        thead th:first-child { text-align: left; border-radius: 8px 0 0 0; }
        thead th:last-child { border-radius: 0 8px 0 0; }
        tbody tr:nth-child(even) { background: #f9fafb; }
        tbody tr:hover { background: #eff6ff; }
        tbody td { padding: 10px 14px; text-align: center; border-bottom: 1px solid #e5e7eb; }
        tbody td:first-child { text-align: left; font-weight: 600; }

        /* Badge camera */
        .cam-chip {
            display: inline-block;
            background: #eef2ff; color: #3730a3;
            padding: 3px 9px; border-radius: 999px;
            font-size: 12px; margin: 2px;
        }

        /* Bar inline */
        .bar-wrap { display: flex; align-items: center; gap: 8px; }
        .bar-bg { flex: 1; background: #e5e7eb; border-radius: 999px; height: 8px; overflow: hidden; }
        .bar-fill { height: 100%; border-radius: 999px; background: #2563eb; }
        .bar-num { font-size: 13px; font-weight: 600; min-width: 36px; text-align: right; }

        /* Peak badge */
        .peak { 
            display: inline-block;
            background: #fef3c7; color: #92400e;
            padding: 3px 9px; border-radius: 6px; font-size: 12px; font-weight: 600;
        }

        /* Tổng row */
        .total-row td { font-weight: 700; background: #e5e7eb !important; }

        /* Camera summary */
        .cam-summary {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
            gap: 14px;
        }
        .cam-card {
            border: 1px solid #e5e7eb; border-radius: 12px;
            padding: 16px 18px;
        }
        .cam-card .cam-name { font-size: 15px; font-weight: 700; color: #2563eb; margin-bottom: 6px; }
        .cam-card .cam-count { font-size: 28px; font-weight: 700; }
        .cam-card .cam-label { font-size: 12px; color: #9ca3af; }
    </style>
</head>
<body>

<header>
    <h1>📊 Thống kê vi phạm</h1>
    <a class="back" href="/">← Quay lại trang chủ</a>
</header>

<div class="container">

    <!-- KPI -->
    <div class="kpi-row">
        <div class="kpi">
            <div class="label">Tổng số ảnh</div>
            <div class="value">{{ total_images }}</div>
            <div class="sub">toàn bộ hệ thống</div>
        </div>
        <div class="kpi">
            <div class="label">Số ngày có dữ liệu</div>
            <div class="value">{{ total_days }}</div>
            <div class="sub">ngày ghi nhận</div>
        </div>
        <div class="kpi">
            <div class="label">Số camera</div>
            <div class="value">{{ all_cams | length }}</div>
            <div class="sub">{% for cam in all_cams %}{{ CAMERA_NAMES.get(cam, cam) }}{% if not loop.last %}, {% endif %}{% endfor %}</div>
        </div>
        {% if total_days > 0 %}
        <div class="kpi">
            <div class="label">TB ảnh / ngày</div>
            <div class="value">{{ (total_images / total_days) | round(1) }}</div>
            <div class="sub">ảnh trung bình</div>
        </div>
        {% endif %}
    </div>

    <!-- Tổng theo camera -->
    <div class="section">
        <h2>📷 Tổng số ảnh theo camera</h2>
        <div class="cam-summary">
            {% for cam in all_cams %}
            <div class="cam-card">
                <div class="cam-name">{{ CAMERA_NAMES.get(cam, cam) }}</div>
                <div class="cam-count">{{ cam_total[cam] }}</div>
                <div class="cam-label">ảnh vi phạm</div>
            </div>
            {% endfor %}
        </div>
    </div>

    <!-- Bảng chi tiết theo ngày -->
    <div class="section">
        <h2>📅 Bảng chi tiết theo ngày</h2>
        <table>
            <thead>
                <tr>
                    <th>Ngày</th>
                    {% for cam in all_cams %}
                    <th>{{ CAMERA_NAMES.get(cam, cam) }}</th>
                    {% endfor %}
                    <th>Giờ cao điểm</th>
                    <th>Tổng</th>
                    <th>Biểu đồ</th>
                </tr>
            </thead>
            <tbody>
                {% set max_total = day_stats | map(attribute='total') | max %}
                {% for day in day_stats %}
                <tr>
                    <td>
                        <a href="/day/{{ day.date }}" style="color:#2563eb;text-decoration:none;">
                            📅 {{ day.date }}
                        </a>
                    </td>
                    {% for cam in all_cams %}
                    <td>
                        {% if day.cameras.get(cam, 0) > 0 %}
                            <span class="cam-chip">{{ day.cameras.get(cam, 0) }}</span>
                        {% else %}
                            <span style="color:#d1d5db;">–</span>
                        {% endif %}
                    </td>
                    {% endfor %}
                    <td>
                        <span class="peak">{{ day.peak_hour }} ({{ day.peak_count }} ảnh)</span>
                    </td>
                    <td><strong>{{ day.total }}</strong></td>
                    <td>
                        <div class="bar-wrap">
                            <div class="bar-bg">
                                <div class="bar-fill" style="width: {{ (day.total / max_total * 100) | round }}%"></div>
                            </div>
                        </div>
                    </td>
                </tr>
                {% endfor %}

                <!-- Tổng cộng -->
                <tr class="total-row">
                    <td>Tổng cộng</td>
                    {% for cam in all_cams %}
                    <td>{{ cam_total[cam] }}</td>
                    {% endfor %}
                    <td>–</td>
                    <td>{{ total_images }}</td>
                    <td></td>
                </tr>
            </tbody>
        </table>
    </div>

</div>

<footer style="background: #ffffff; padding: 30px 0; border-top: 1px solid #e5e7eb; margin-top: 50px;">
    <div style="max-width: 1100px; margin: 0 auto; padding: 0 32px; display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 20px;">
        
        <div style="flex: 1; min-width: 300px;">
            <h3 style="margin: 0 0 10px 0; font-size: 18px; color: #111827; text-transform: uppercase; letter-spacing: 1px;">Thông tin liên hệ</h3>
            <p style="margin: 5px 0; color: #4b5563; font-size: 14px; line-height: 1.6;">
                <strong>Địa chỉ:</strong> 32/19 Đường 494, phường Tăng Nhơn Phú, TP Hồ Chí Minh
            </p>
        </div>

        <div style="flex: 1; min-width: 250px; text-align: right;">
            <p style="margin: 5px 0; color: #4b5563; font-size: 14px;">
                <strong>Email:</strong> ipx@gmail.com
            </p>
            <p style="margin: 5px 0; color: #1d4ed8; font-size: 16px; font-weight: bold;">
                Hotline: 0983 432 188
            </p>
            <p style="margin: 5px 0; color: #4b5563; font-size: 14px;">
                Hỗ trợ kỹ thuật: 0982 805 485
            </p>
        </div>

    </div>
    
    <div style="text-align: center; margin-top: 25px; padding-top: 15px; border-top: 1px solid #f3f4f6; color: #9ca3af; font-size: 12px;">
        &copy; 2024 Bản quyền thuộc về Công ty cổ phần công nghệ IPX.
    </div>
</footer>
</body>
</html>
"""


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
